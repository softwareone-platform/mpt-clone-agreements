import logging
import os
import sys
import time
from pathlib import Path
from dotenv import load_dotenv
from urllib.parse import urljoin
from typing import Dict, List, Any, Optional, Tuple
import argparse
import json
from openpyxl import load_workbook

# Add the script's directory to the path to allow importing utils
script_dir = Path(__file__).parent
if str(script_dir) not in sys.path:
    sys.path.insert(0, str(script_dir))

from utils import (
    validate_agreement_id,
    setup_logging,
    validate_agreement_and_tokens,
    create_http_client,
    make_request_with_retry
)


# Load environment variables from ~/.mpt-clone-agreement
ENV_PATH = Path.home() / '.mpt-clone-agreement'
load_dotenv(ENV_PATH)


class ConfigurationManager:
    REQUIRED_VARS = {
        'OPS_TOKEN': 'API Operations Token',
        'VENDOR_TOKEN': 'Vendor API Token',
        'API_URL': 'API Base URL',
    }
    
    OPTIONAL_CSP_VARS = {
        'CSP_URL_TUNNEL': 'CSP Tunnel URL',
        'CSP_TOKEN': 'CSP Token',
    }

    @classmethod
    def load_config(cls, require_csp: bool = False) -> Dict[str, str]:
        """
        Load and validate environment variables.
        
        Args:
            require_csp: If True, also require CSP_URL_TUNNEL and CSP_TOKEN
            
        Returns:
            Dictionary of configuration values
        """
        # Always load required vars
        required_vars = cls.REQUIRED_VARS.copy()
        
        # Conditionally add CSP vars if needed
        if require_csp:
            required_vars.update(cls.OPTIONAL_CSP_VARS)
        
        config = {key: os.getenv(key) for key in required_vars}
        missing_vars = [f"{key} ({desc})" for key, desc in required_vars.items()
                       if not config.get(key)]
        if missing_vars:
            raise EnvironmentError(
                f"Missing required environment variables:\n"
                f"{chr(10).join(missing_vars)}\n"
                f"Please ensure these are set in {ENV_PATH}"
            )
        return config


def parse_arguments() -> argparse.Namespace:
    """
    Parse command line arguments.
    
    Returns:
        argparse.Namespace: Parsed command line arguments
    """
    parser = argparse.ArgumentParser(description='Create new agreement from new_agreement_object.json')
    parser.add_argument(
        '--agreement-id',
        type=str,
        required=True,
        dest='agreement_id',
        help='Agreement ID to locate the folder (e.g., AGR-1234-5678-9012)'
    )
    parser.add_argument(
        '--debug',
        action='store_true',
        help='Enable debug level logging (more verbose output)'
    )
    parser.add_argument(
        '--microsoft-sync',
        action='store_true',
        dest='microsoft_sync',
        help='Trigger Microsoft platform synchronization (default: create subscriptions from Excel)'
    )
    parser.add_argument(
        '--keep-purchase-price',
        action='store_true',
        dest='keep_purchase_price',
        help='Keep the price node in line items when creating subscriptions'
    )
    args = parser.parse_args()
    
    validate_agreement_id(args.agreement_id)
    
    return args


def load_new_agreement_object(
    agreement_id: str,
    logger: logging.Logger
) -> Optional[Dict[str, Any]]:
    """Load new_agreement_object.json from output folder."""
    agreement_file = Path("output") / agreement_id / "new_agreement_object.json"
    
    if not agreement_file.exists():
        logger.error(f"File not found: {agreement_file}")
        return None
    
    try:
        with open(agreement_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
        logger.info(f"Successfully loaded new_agreement_object.json from {agreement_file}")
        return data
    except json.JSONDecodeError as e:
        logger.error(f"Invalid JSON in {agreement_file}: {str(e)}")
        return None
    except Exception as e:
        logger.error(f"Error reading {agreement_file}: {str(e)}")
        return None


def load_authorization_object(
    agreement_id: str,
    logger: logging.Logger
) -> Optional[Dict[str, Any]]:
    """Load authorization.json from output folder."""
    authorization_file = Path("output") / agreement_id / "authorization.json"
    
    if not authorization_file.exists():
        logger.error(f"File not found: {authorization_file}")
        return None
    
    try:
        with open(authorization_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
        logger.info(f"Successfully loaded authorization.json from {authorization_file}")
        return data
    except json.JSONDecodeError as e:
        logger.error(f"Invalid JSON in {authorization_file}: {str(e)}")
        return None
    except Exception as e:
        logger.error(f"Error reading {authorization_file}: {str(e)}")
        return None


def remove_fields_for_post(
    agreement: Dict[str, Any],
    logger: logging.Logger
) -> Dict[str, Any]:
    """Create a copy of agreement without externalIds.vendor and parameters.fulfillment."""
    agreement_copy = json.loads(json.dumps(agreement))
    
    if 'externalIds' in agreement_copy and 'vendor' in agreement_copy['externalIds']:
        del agreement_copy['externalIds']['vendor']
        logger.debug("Removed externalIds.vendor from POST payload")
    
    if 'parameters' in agreement_copy and 'fulfillment' in agreement_copy['parameters']:
        del agreement_copy['parameters']['fulfillment']
        logger.debug("Removed parameters.fulfillment from POST payload")

    if 'certificates' in agreement_copy:
        del agreement_copy['certificates']
        logger.debug("Removed certificates from POST payload")

    return agreement_copy


def create_agreement(
    base_url: str,
    ops_token: str,
    payload: Dict[str, Any],
    logger: logging.Logger
) -> Optional[str]:
    """Create agreement via POST request using OPS token."""
    url = urljoin(base_url, '/public/v1/commerce/agreements')
    
    logger.info(f"Creating agreement via POST to {url}")
    logger.debug(f"POST payload: {json.dumps(payload, indent=2)}")
    
    with create_http_client(ops_token, 'Agreement Clone Creator') as client:
        result = make_request_with_retry(
            client=client,
            method='POST',
            url=url,
            logger=logger,
            parse_json=True,
            json=payload
        )
    
    if not result:
        logger.error("Failed to create agreement")
        return None
    
    agreement_id = result.get('id')
    if agreement_id:
        logger.info(f"Successfully created agreement: {agreement_id}")
        return agreement_id
    else:
        logger.error(f"POST succeeded but no agreement ID in response: {result}")
        return None


def perform_platform_sync(
    base_url: str,
    auth_id: str,
    tenant_id: str,
    auth_token: str,
    logger: logging.Logger
) -> Tuple[bool, int, str]:
    """
    Call platform sync endpoint for a specific authorization and tenant.
    
    Args:
        base_url: API base URL
        auth_id: Authorization ID
        tenant_id: Tenant ID (client ID)
        auth_token: Authentication token
        logger: Logger instance
        
    Returns:
        Tuple of (success, status_code, response_text)
    """
    try:
        synchronization_key = str(int(time.time()))
        url = urljoin(
            base_url,
            f'/v1/maintenance/authorizations/{auth_id}/customers/{tenant_id}/sync?synchronizationKey={synchronization_key}'
        )
        
        logger.info(f"Triggering platform sync for authorization {auth_id}, tenant {tenant_id}")
        logger.debug(f"Platform sync URL: {url}")
        logger.debug(f"Authorization ID: {auth_id}")
        logger.debug(f"Tenant ID: {tenant_id}")
        logger.debug(f"Synchronization Key: {synchronization_key}")
        
        with create_http_client(auth_token, 'Agreement Clone Creator') as client:
            resp = make_request_with_retry(
                client=client,
                method='POST',
                url=url,
                logger=logger,
                parse_json=False,
                json={}
            )
        
        if not resp:
            logger.error(f"Platform sync failed for {auth_id}/{tenant_id}")
            return False, 0, "Request failed"
        
        logger.info(f"Platform sync response: Status {resp.status_code} | Auth: {auth_id} | Tenant: {tenant_id}")
        
        response_text = resp.text
        if response_text:
            logger.debug(f"Platform sync response body: {response_text}")
        
        success = 200 <= resp.status_code < 300
        if success:
            logger.info(f"Platform sync completed successfully for {auth_id}/{tenant_id}")
        else:
            logger.warning(f"Platform sync returned non-success status {resp.status_code} for {auth_id}/{tenant_id}")
        
        return success, resp.status_code, response_text
        
    except Exception as e:
        logger.error(f"Platform sync error for {auth_id}/{tenant_id}: {str(e)}")
        return False, 0, str(e)


def get_agreement_details(
    base_url: str,
    ops_token: str,
    agreement_id: str,
    logger: logging.Logger
) -> Optional[Dict[str, Any]]:
    """Fetch agreement details from API."""
    url = urljoin(base_url, f'/public/v1/commerce/agreements/{agreement_id}')
    
    logger.debug(f"Fetching agreement details from {url}")
    
    with create_http_client(ops_token, 'Agreement Clone Creator') as client:
        agreement = make_request_with_retry(
            client=client,
            method='GET',
            url=url,
            logger=logger,
            parse_json=True
        )
    
    if agreement:
        logger.debug(f"Successfully fetched agreement details for {agreement_id}")
    else:
        logger.error(f"Failed to fetch agreement {agreement_id}")
    
    return agreement


def read_subscription_ids_from_excel(
    agreement_id: str,
    logger: logging.Logger
) -> List[str]:
    """Read subscription IDs from Excel file (column 'ID')."""
    excel_file = Path("output") / agreement_id / "subscriptions.xlsx"
    
    if not excel_file.exists():
        logger.error(f"Excel file not found: {excel_file}")
        return []
    
    try:
        wb = load_workbook(excel_file, data_only=True)
        ws = wb.active
        
        id_column = None
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            if header and str(header).strip() == 'ID':
                id_column = col
                break
        
        if id_column is None:
            logger.error("'ID' column not found in Excel file")
            return []
        
        subscription_ids = []
        seen_ids = set()
        
        for row in range(2, ws.max_row + 1):
            sub_id = ws.cell(row=row, column=id_column).value
            if sub_id:
                sub_id_str = str(sub_id).strip()
                if sub_id_str not in seen_ids:
                    subscription_ids.append(sub_id_str)
                    seen_ids.add(sub_id_str)
        
        logger.info(f"Read {len(subscription_ids)} unique subscription IDs from Excel file (deduplicated from {ws.max_row - 1} rows)")
        return subscription_ids
    except Exception as e:
        logger.error(f"Error reading Excel file {excel_file}: {str(e)}")
        return []


def load_subscription_json(
    agreement_id: str,
    subscription_id: str,
    logger: logging.Logger
) -> Optional[Dict[str, Any]]:
    """Load subscription JSON file that was dumped by dump_agreement.py."""
    json_file = Path("output") / agreement_id / f"{subscription_id}.json"
    
    if not json_file.exists():
        logger.warning(f"Subscription JSON file not found: {json_file}")
        return None
    
    try:
        with open(json_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
        logger.debug(f"Successfully loaded subscription JSON from {json_file}")
        return data
    except json.JSONDecodeError as e:
        logger.error(f"Invalid JSON in {json_file}: {str(e)}")
        return None
    except Exception as e:
        logger.error(f"Error reading {json_file}: {str(e)}")
        return None


def filter_subscription_for_post(
    subscription_data: Dict[str, Any],
    logger: logging.Logger,
    keep_purchase_price: bool = False
) -> Dict[str, Any]:
    """
    Filter subscription data to only include required first-level properties for POST.
    
    Required properties: agreement, autoRenew, commitmentDate, externalIds, lines, 
    name, parameters, startDate, template
    
    For lines: only item.id and quantity are kept for each line.
    If keep_purchase_price is True, the price node is also kept in each line.
    """
    allowed_properties = {
        'agreement',
        'autoRenew',
        'commitmentDate',
        'externalIds',
        'lines',
        'name',
        'parameters',
        'startDate',
        'template'
    }
    
    filtered_data = {}
    for key in allowed_properties:
        if key in subscription_data:
            if key == 'lines':
                lines = subscription_data[key]
                sanitized_lines = []
                for line in lines:
                    sanitized_line = {
                        'item': {'id': line['item']['id']},
                        'quantity': line['quantity']
                    }
                    if keep_purchase_price and 'price' in line:
                        sanitized_line['price'] = line['price']
                    sanitized_lines.append(sanitized_line)
                filtered_data[key] = sanitized_lines
                if keep_purchase_price:
                    logger.debug(f"Sanitized {len(sanitized_lines)} line(s), keeping item.id, quantity, and price")
                else:
                    logger.debug(f"Sanitized {len(sanitized_lines)} line(s), keeping only item.id and quantity")
            else:
                filtered_data[key] = subscription_data[key]
        else:
            logger.debug(f"Property '{key}' not found in subscription data, will be omitted")
    
    removed_properties = set(subscription_data.keys()) - allowed_properties
    if removed_properties:
        logger.debug(f"Removed properties from subscription data: {sorted(removed_properties)}")
    
    return filtered_data


def create_subscription(
    base_url: str,
    vendor_token: str,
    subscription_data: Dict[str, Any],
    logger: logging.Logger,
    keep_purchase_price: bool = False
) -> Optional[str]:
    """Create subscription via POST request using VENDOR token."""
    url = urljoin(base_url, '/public/v1/commerce/subscriptions')
    
    filtered_data = filter_subscription_for_post(subscription_data, logger, keep_purchase_price)
    
    logger.info(f"Creating subscription via POST to {url}")
    logger.debug(f"POST payload: {json.dumps(filtered_data, indent=2)}")
    
    with create_http_client(vendor_token, 'Microsoft Subscription Creator') as client:
        result = make_request_with_retry(
            client=client,
            method='POST',
            url=url,
            logger=logger,
            parse_json=True,
            json=filtered_data
        )
    
    if not result:
        logger.error("Failed to create subscription")
        return None
    
    subscription_id = result.get('id')
    if subscription_id:
        logger.info(f"Successfully created subscription: {subscription_id}")
        return subscription_id
    else:
        logger.error(f"POST succeeded but no subscription ID in response: {result}")
        return None


def update_agreement_certificates(
    base_url: str,
    vendor_token: str,
    agreement_id: str,
    certificates: List[Dict[str, Any]],
    logger: logging.Logger
) -> bool:
    """
    Update certificates in the agreement via PUT using VENDOR token.
    
    Args:
        base_url: API base URL
        vendor_token: Vendor token for authentication
        agreement_id: Agreement ID to update
        certificates: List of certificate objects (will extract only IDs)
        logger: Logger instance
        
    Returns:
        True if successful, False otherwise
    """
    url = urljoin(base_url, f'/public/v1/commerce/agreements/{agreement_id}')
    
    certificate_ids = []
    for cert in certificates:
        cert_id = cert.get('id') if isinstance(cert, dict) else None
        if cert_id:
            certificate_ids.append({'id': cert_id})
    
    if not certificate_ids:
        logger.warning("No valid certificate IDs found in certificates array")
        return False
    
    payload = {
        'certificates': certificate_ids
    }
    
    logger.info(f"Updating certificates via PUT to {url}")
    logger.debug(f"PUT payload: {json.dumps(payload, indent=2)}")
    
    with create_http_client(vendor_token, 'Agreement Clone Creator') as client:
        result = make_request_with_retry(
            client=client,
            method='PUT',
            url=url,
            logger=logger,
            parse_json=False,
            json=payload
        )
    
    if result and 200 <= result.status_code < 300:
        logger.info(f"Successfully updated certificates for agreement {agreement_id}")
        return True
    else:
        logger.error(f"Failed to update certificates for agreement {agreement_id}")
        return False


def update_agreement_field(
    base_url: str,
    vendor_token: str,
    agreement_id: str,
    field_path: str,
    value: Any,
    logger: logging.Logger
) -> bool:
    """Update a specific field in the agreement via PUT using VENDOR token."""
    url = urljoin(base_url, f'/public/v1/commerce/agreements/{agreement_id}')
    
    payload = {'id': agreement_id}
    
    parts = field_path.split('.')
    current = payload
    for part in parts[:-1]:
        if part not in current:
            current[part] = {}
        current = current[part]
    current[parts[-1]] = value
    
    logger.info(f"Updating {field_path} via PUT to {url}")
    logger.debug(f"PUT payload: {json.dumps(payload, indent=2)}")
    
    with create_http_client(vendor_token, 'Microsoft Agreement Updater') as client:
        result = make_request_with_retry(
            client=client,
            method='PUT',
            url=url,
            logger=logger,
            parse_json=False,
            json=payload
        )
    
    if result and 200 <= result.status_code < 300:
        logger.info(f"Successfully updated {field_path} for agreement {agreement_id}")
        return True
    else:
        logger.error(f"Failed to update {field_path} for agreement {agreement_id}")
        return False


def main():
    args = parse_arguments()
    script_name = Path(__file__).stem
    logger = setup_logging(script_name, args.debug, args.agreement_id)
    
    try:
        logger.debug("Starting agreement creation in debug mode")
        # Only require CSP config if --microsoft-sync flag is set
        config = ConfigurationManager.load_config(require_csp=args.microsoft_sync)
        
        # Validate source agreement and tokens before proceeding
        validate_agreement_and_tokens(
            args.agreement_id,
            config['API_URL'],
            config['OPS_TOKEN'],
            config['VENDOR_TOKEN'],
            logger
        )
        
        # Load new_agreement_object.json
        logger.debug(f"Loading new_agreement_object.json for agreement {args.agreement_id}")
        agreement_data = load_new_agreement_object(args.agreement_id, logger)
        if not agreement_data:
            logger.error("Failed to load new_agreement_object.json, stopping execution")
            return
        
        logger.debug("Preparing POST payload (removing externalIds.vendor and parameters.fulfillment)")
        post_payload = remove_fields_for_post(agreement_data, logger)
        
        external_ids_vendor = agreement_data.get('externalIds', {}).get('vendor')
        parameters_fulfillment = agreement_data.get('parameters', {}).get('fulfillment')
        template_id = agreement_data.get('template', {}).get('id')
        certificates = agreement_data.get('certificates', [])
        
        # Create agreement via POST
        logger.info("Step 1: Creating agreement via POST (OPS_TOKEN)")
        new_agreement_id = create_agreement(
            config['API_URL'],
            config['OPS_TOKEN'],
            post_payload,
            logger
        )
        
        if not new_agreement_id:
            logger.error("Failed to create agreement, stopping execution")
            return
        
        logger.info(f"Agreement created successfully: {new_agreement_id}")
        
        if parameters_fulfillment:
            logger.info("Step 2: Updating parameters.fulfillment via PUT (VENDOR_TOKEN)")
            success = update_agreement_field(
                config['API_URL'],
                config['VENDOR_TOKEN'],
                new_agreement_id,
                'parameters.fulfillment',
                parameters_fulfillment,
                logger
            )
            if not success:
                logger.warning("Failed to update parameters.fulfillment, continuing...")
        else:
            logger.warning("No parameters.fulfillment found in original agreement, skipping PUT")
        
        if external_ids_vendor:
            logger.info("Step 3: Updating externalIds.vendor via PUT (VENDOR_TOKEN)")
            success = update_agreement_field(
                config['API_URL'],
                config['VENDOR_TOKEN'],
                new_agreement_id,
                'externalIds.vendor',
                external_ids_vendor,
                logger
            )
            if not success:
                logger.warning("Failed to update externalIds.vendor, continuing...")
        else:
            logger.warning("No externalIds.vendor found in original agreement, skipping PUT")
        
        if template_id:
            logger.info("Step 4: Updating template.id via PUT (VENDOR_TOKEN)")
            success = update_agreement_field(
                config['API_URL'],
                config['VENDOR_TOKEN'],
                new_agreement_id,
                'template.id',
                template_id,
                logger
            )
            if not success:
                logger.warning("Failed to update template.id, continuing...")
        else:
            logger.warning("No template.id found in original agreement, skipping PUT")
        
        if certificates:
            logger.info("Step 5: Updating certificates via PUT (VENDOR_TOKEN)")
            success = update_agreement_certificates(
                config['API_URL'],
                config['VENDOR_TOKEN'],
                new_agreement_id,
                certificates,
                logger
            )
            if not success:
                logger.warning("Failed to update certificates, continuing...")
        else:
            logger.warning("No certificates found in original agreement, skipping PUT")
        
        if args.microsoft_sync:
            logger.info("Step 6: Triggering platform synchronization")
            
            if 'CSP_URL_TUNNEL' not in config or 'CSP_TOKEN' not in config:
                logger.error("CSP_URL_TUNNEL and CSP_TOKEN are required when --microsoft-sync is used")
                return
            
            authorization_data = load_authorization_object(args.agreement_id, logger)
            if not authorization_data:
                logger.warning("Could not load authorization.json, skipping platform sync")
            else:
                auth_id = authorization_data.get('externalIds', {}).get('operations')
                tenant_id = external_ids_vendor
                
                if auth_id and tenant_id:
                    sync_success, sync_status, sync_response = perform_platform_sync(
                        config['CSP_URL_TUNNEL'],
                        auth_id,
                        tenant_id,
                        config['CSP_TOKEN'],
                        logger
                    )
                    if sync_success:
                        logger.info("Platform synchronization completed successfully")
                    else:
                        logger.warning(f"Platform synchronization returned status {sync_status}: {sync_response}")
                else:
                    logger.warning(f"Could not extract authorization ID or tenant ID. Auth ID: {auth_id}, Tenant ID: {tenant_id}")
        else:
            logger.info("Step 6: Creating subscriptions from Excel and JSON files")
            
            subscription_ids = read_subscription_ids_from_excel(args.agreement_id, logger)
            if not subscription_ids:
                logger.warning("No subscription IDs found in Excel file, skipping subscription creation")
            else:
                created_count = 0
                failed_count = 0
                
                for sub_id in subscription_ids:
                    logger.info(f"Processing subscription ID: {sub_id}")
                    
                    subscription_data = load_subscription_json(args.agreement_id, sub_id, logger)
                    if not subscription_data:
                        logger.warning(f"Could not load subscription JSON for {sub_id}, skipping")
                        failed_count += 1
                        continue
                    
                    if 'agreement' not in subscription_data:
                        subscription_data['agreement'] = {}
                    subscription_data['agreement']['id'] = new_agreement_id
                    
                    created_sub_id = create_subscription(
                        config['API_URL'],
                        config['VENDOR_TOKEN'],
                        subscription_data,
                        logger,
                        args.keep_purchase_price
                    )
                    
                    if created_sub_id:
                        created_count += 1
                        logger.info(f"Successfully created subscription {created_sub_id} (original: {sub_id})")
                    else:
                        failed_count += 1
                        logger.error(f"Failed to create subscription {sub_id}")
                
                logger.info(f"Subscription creation summary: {created_count} created, {failed_count} failed out of {len(subscription_ids)} total")
        
        logger.info("Step 7: Fetching final agreement details")
        final_agreement = get_agreement_details(
            config['API_URL'],
            config['OPS_TOKEN'],
            new_agreement_id,
            logger
        )
        
        if final_agreement:
            output_dir = Path("output") / args.agreement_id
            final_agreement_json_path = output_dir / 'final_agreement.json'
            with open(final_agreement_json_path, 'w', encoding='utf-8') as f:
                json.dump(final_agreement, f, indent=2, ensure_ascii=False)
            logger.info(f"Final agreement saved to {final_agreement_json_path}")
        else:
            logger.warning("Could not fetch final agreement details, skipping save")
        
        logger.info(f"Agreement creation process completed for {new_agreement_id}")
        logger.info("All steps completed successfully")
        
    except Exception as e:
        logger.error(f"Error in main execution: {str(e)}")
        raise


if __name__ == "__main__":
    main()

