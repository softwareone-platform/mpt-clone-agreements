import logging
import os
import sys
from pathlib import Path
from dotenv import load_dotenv
from datetime import datetime, timezone
from urllib.parse import urljoin
from typing import Dict, List, Any, Optional
import argparse
import json
from openpyxl import load_workbook

# Add the script's directory to the path to allow importing utils
script_dir = Path(__file__).parent
if str(script_dir) not in sys.path:
    sys.path.insert(0, str(script_dir))

from utils import (
    validate_agreement_id,
    setup_logging,
    validate_agreement_and_tokens,
    create_http_client,
    make_request_with_retry,
    calculate_unit_sp
)


# Load environment variables from ~/.mpt-clone-agreement
ENV_PATH = Path.home() / '.mpt-clone-agreement'
load_dotenv(ENV_PATH)


class ConfigurationManager:
    REQUIRED_VARS = {
        'OPS_TOKEN': 'API Operations Token',
        'VENDOR_TOKEN': 'Vendor API Token',
        'API_URL': 'API Base URL',
    }

    @classmethod
    def load_config(cls) -> Dict[str, str]:
        """Load and validate environment variables."""
        config = {key: os.getenv(key) for key in cls.REQUIRED_VARS}
        missing_vars = [f"{key} ({desc})" for key, desc in cls.REQUIRED_VARS.items()
                        if not config.get(key)]
        if missing_vars:
            raise EnvironmentError(
                f"Missing required environment variables:\n"
                f"{chr(10).join(missing_vars)}\n"
                f"Please ensure these are set in {ENV_PATH}"
            )
        return config


def parse_arguments() -> argparse.Namespace:
    """
    Parse command line arguments.
    
    Returns:
        argparse.Namespace: Parsed command line arguments
    """
    parser = argparse.ArgumentParser(description='Update subscription markups from Excel file')
    parser.add_argument(
        '--agreement-id',
        type=str,
        required=True,
        dest='agreement_id',
        help='Agreement ID to locate the folder (e.g., AGR-1234-5678-9012)'
    )
    parser.add_argument(
        '--no-dry-run',
        action='store_true',
        default=True,
        help='Disable dry-run mode and make actual changes'
    )
    parser.add_argument(
        '--keep-purchase-price',
        action='store_true',
        help='Keep purchase price by using unitPP and calculating unitSP from unitPP and markup (falls back to markup if unitPP is not available)'
    )
    parser.add_argument(
        '--debug',
        action='store_true',
        help='Enable debug level logging (more verbose output)'
    )
    args = parser.parse_args()
    
    validate_agreement_id(args.agreement_id)
    
    return args


def load_final_agreement(
    agreement_id: str,
    logger: logging.Logger
) -> Optional[Dict[str, Any]]:
    """Load final_agreement.json from output folder."""
    agreement_file = Path("output") / agreement_id / "final_agreement.json"
    
    if not agreement_file.exists():
        logger.error(f"File not found: {agreement_file}")
        return None
    
    try:
        with open(agreement_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
        logger.info(f"Successfully loaded final_agreement.json from {agreement_file}")
        return data
    except json.JSONDecodeError as e:
        logger.error(f"Invalid JSON in {agreement_file}: {str(e)}")
        return None
    except Exception as e:
        logger.error(f"Error reading {agreement_file}: {str(e)}")
        return None


def read_excel_subscriptions(
    agreement_id: str,
    logger: logging.Logger
) -> Dict[str, List[Dict[str, Any]]]:
    """
    Read subscriptions from Excel file and group by Vendor Sub ID.
    
    Returns:
        Dictionary mapping Vendor Sub ID to list of items with their markups
    """
    excel_file = Path("output") / agreement_id / "subscriptions.xlsx"
    
    if not excel_file.exists():
        logger.error(f"File not found: {excel_file}")
        return {}
    
    try:
        wb = load_workbook(excel_file, data_only=True)
        ws = wb.active
        
        headers = {}
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            if header:
                headers[header] = col
        
        required_cols = ['ID', 'Vendor Sub ID', 'Item ID', 'Markup']
        missing_cols = [col for col in required_cols if col not in headers]
        if missing_cols:
            logger.error(f"Excel file is missing required columns: {missing_cols}")
            return {}
        
        unit_pp_col = 27
        unit_pp_header = 'Unit PP'
        if unit_pp_header in headers:
            has_unit_pp = True
        else:
            header_at_col_27 = ws.cell(row=1, column=unit_pp_col).value
            if header_at_col_27 and str(header_at_col_27).strip() == unit_pp_header:
                headers[unit_pp_header] = unit_pp_col
                has_unit_pp = True
            else:
                if unit_pp_col <= ws.max_column:
                    headers[unit_pp_header] = unit_pp_col
                    has_unit_pp = True
                    logger.debug(f"Using column {unit_pp_col} (AA) for Unit PP (header: '{header_at_col_27}')")
                else:
                    has_unit_pp = False
                    logger.warning(f"Column {unit_pp_col} (AA) not found. Will try to use API value if needed.")
        
        subscriptions_map = {}
        total_rows = 0
        
        for row in range(2, ws.max_row + 1):
            sub_id = ws.cell(row=row, column=headers['ID']).value
            vendor_sub_id = ws.cell(row=row, column=headers['Vendor Sub ID']).value
            item_id = ws.cell(row=row, column=headers['Item ID']).value
            markup = ws.cell(row=row, column=headers['Markup']).value
            
            unit_pp = None
            if has_unit_pp:
                unit_pp_value = ws.cell(row=row, column=headers[unit_pp_header]).value
                if unit_pp_value is not None:
                    try:
                        unit_pp = float(unit_pp_value)
                        if unit_pp == 0.0:
                            unit_pp = None
                    except (ValueError, TypeError):
                        logger.debug(f"Row {row}: Could not convert Unit PP value '{unit_pp_value}' to float")
            
            if sub_id and vendor_sub_id and item_id and markup is not None:
                vendor_sub_id_str = str(vendor_sub_id) if vendor_sub_id else None
                if vendor_sub_id_str:
                    if vendor_sub_id_str not in subscriptions_map:
                        subscriptions_map[vendor_sub_id_str] = {
                            'subscription_id': str(sub_id),
                            'items': []
                        }
                    subscriptions_map[vendor_sub_id_str]['items'].append({
                        'Item ID': str(item_id),
                        'Markup': float(markup) if markup is not None else None,
                        'Unit PP': unit_pp
                    })
                    total_rows += 1
        
        logger.info(f"Read {total_rows} rows from Excel file, grouped into {len(subscriptions_map)} unique subscriptions")
        return subscriptions_map
    except Exception as e:
        logger.error(f"Error reading Excel file {excel_file}: {str(e)}")
        return {}


def fetch_agreement_subscriptions(
    base_url: str,
    ops_token: str,
    agreement_id: str,
    logger: logging.Logger
) -> List[Dict[str, Any]]:
    """Fetch all subscriptions for the agreement."""
    items = []
    page = None
    limit = 1000
    offset = 0
    
    try:
        current_date = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%S.000Z")
    except Exception:
        current_date = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%S.000Z")
    
    url_base = urljoin(
        base_url,
        f'/public/v1/commerce/subscriptions?and(lt(audit.created.at,{current_date}),eq(agreement.id,{agreement_id}),eq(status,active))'
        '&select=agreement,lines,agreement.authorization.externalIds,agreement.listing.priceList,agreement.parameters,agreement.certificates,licensee,buyer,seller,audit,-agreement.subscriptions,-agreement.lines'
        f'&order=-audit.created.at'
    )
    
    with create_http_client(ops_token, 'Clone Agreement Markup Updater') as client:
        while True:
            url = f"{url_base}&offset={offset}&limit={limit}"
            page = make_request_with_retry(
                client=client,
                method='GET',
                url=url,
                logger=logger,
                parse_json=True
            )
            
            if not page:
                logger.error(f"Failed to fetch subscriptions page at offset {offset}")
                break
            
            page_content = page.get('data', [])
            items.extend(page_content)
            
            meta = page.get('$meta', {}).get('pagination', {})
            offset = meta.get('offset', 0)
            limit = meta.get('limit', 0)
            total = meta.get('total', 0)
            
            if offset + limit >= total:
                break
            
            offset += limit
            logger.debug(f"Fetching next page: offset={offset}")
    
    logger.info(f"Fetched {len(items)} subscriptions from agreement {agreement_id}")
    return items


def update_subscription_markup(
    base_url: str,
    ops_token: str,
    subscription_id: str,
    lines_data: List[Dict[str, Any]],
    logger: logging.Logger,
    dry_run: bool = True,
    keep_purchase_price: bool = False
) -> bool:
    """
    Update subscription markups for multiple lines using PUT request.
    
    Args:
        base_url: API base URL
        ops_token: OPS token for authentication
        subscription_id: Subscription ID
        lines_data: List of line data dictionaries, each containing:
            - line_id: Line ID
            - markup: Markup value
            - item_id: Item ID
            - terms: Terms dictionary
            - quantity: Quantity
            - quantity_not_applicable: Boolean
        logger: Logger instance
        dry_run: If True, only log what would be done
    """
    url = urljoin(base_url, f'/public/v1/commerce/subscriptions/{subscription_id}')
    
    lines_payload = []
    for line_data in lines_data:
        if keep_purchase_price:
            unit_pp = line_data.get('unit_pp')
            if unit_pp is None or unit_pp == 0.0:
                original_price = line_data.get('price', {})
                api_unit_pp = original_price.get('unitPP')
                if api_unit_pp is not None and api_unit_pp != 0.0:
                    unit_pp = api_unit_pp
            
            markup = line_data['markup']
            
            if unit_pp is not None and unit_pp != 0.0:
                markup_decimal = markup / 100.0
                unit_sp = calculate_unit_sp(unit_pp, markup_decimal)
                price_obj = {
                    "unitPP": unit_pp,
                    "unitSP": unit_sp
                }
            else:
                logger.warning(f"unitPP is 0.0 or not found in Excel or API for line {line_data['line_id']}, using markup instead")
                price_obj = {
                    "markup": markup
                }
        else:
            price_obj = {
                "markup": line_data['markup']
            }
        
        line_payload = {
            "price": price_obj,
            "subscription": {
                "id": subscription_id
            },
            "id": line_data['line_id'],
            "quantity": line_data['quantity'],
            "item": {
                "id": line_data['item_id']
            },
            "quantityNotApplicable": line_data['quantity_not_applicable']
        }
        
        if not keep_purchase_price:
            line_payload["terms"] = line_data['terms']
        
        lines_payload.append(line_payload)
    
    payload = {
        "lines": lines_payload
    }
    
    if dry_run:
        logger.info(f"[DRY RUN] Would update subscription {subscription_id} with {len(lines_payload)} line(s)")
        logger.debug(f"Payload: {json.dumps(payload, indent=2)}")
        return True
    
    logger.info(f"Updating subscription {subscription_id} with {len(lines_payload)} line(s)...")
    logger.debug(f"PUT request to {url}")
    logger.debug(f"Payload: {json.dumps(payload, indent=2)}")
    
    with create_http_client(ops_token, 'Clone Agreement Markup Updater') as client:
        result = make_request_with_retry(
            client=client,
            method='PUT',
            url=url,
            logger=logger,
            parse_json=False,
            json=payload
        )
    
    if result and 200 <= result.status_code < 300:
        logger.info(f"Successfully updated subscription {subscription_id} with {len(lines_payload)} line(s)")
        return True
    else:
        logger.error(f"Failed to update subscription {subscription_id}")
        return False


def main():
    args = parse_arguments()
    script_name = Path(__file__).stem
    logger = setup_logging(script_name, args.debug, args.agreement_id)
    
    try:
        logger.debug("Starting subscription markup update in debug mode")
        config = ConfigurationManager.load_config()
        
        validate_agreement_and_tokens(
            args.agreement_id,
            config['API_URL'],
            config['OPS_TOKEN'],
            config['VENDOR_TOKEN'],
            logger
        )
        
        logger.debug(f"Loading final_agreement.json for agreement {args.agreement_id}")
        final_agreement = load_final_agreement(args.agreement_id, logger)
        if not final_agreement:
            logger.error("Failed to load final_agreement.json, stopping execution")
            return
        
        created_agreement_id = final_agreement.get('id')
        if not created_agreement_id:
            logger.error("final_agreement.json does not contain an 'id' field, stopping execution")
            return
        
        logger.info(f"Using agreement ID from final_agreement.json: {created_agreement_id}")
        
        logger.debug("Reading subscriptions from Excel file")
        excel_subscriptions = read_excel_subscriptions(args.agreement_id, logger)
        if not excel_subscriptions:
            logger.error("No subscriptions found in Excel file, stopping execution")
            return
        
        logger.info(f"Found {len(excel_subscriptions)} unique subscriptions in Excel file")
        
        logger.debug(f"Fetching all subscriptions for agreement {created_agreement_id}")
        agreement_subscriptions = fetch_agreement_subscriptions(
            config['API_URL'],
            config['OPS_TOKEN'],
            created_agreement_id,
            logger
        )
        
        logger.info(f"Found {len(agreement_subscriptions)} subscriptions in agreement")
        
        dry_run = not args.no_dry_run
        if dry_run:
            logger.info("Running in dry-run mode - no changes will be made")
        else:
            logger.info("Running in update mode - changes will be made")
        
        found_count = 0
        updated_count = 0
        not_found_count = 0
        total_lines_updated = 0
        
        for sub in agreement_subscriptions:
            sub_id = sub.get('id')
            external_vendor_id = sub.get('externalIds', {}).get('vendor')
            
            if not external_vendor_id or (isinstance(external_vendor_id, str) and external_vendor_id.strip() == ''):
                logger.debug(f"Subscription {sub_id} has no externalIds.vendor (empty/null), skipping")
                continue
            
            excel_sub_data = excel_subscriptions.get(external_vendor_id)
            if not excel_sub_data:
                logger.debug(f"Subscription {sub_id} (externalId.vendor: {external_vendor_id}) not found in Excel")
                not_found_count += 1
                continue
            
            found_count += 1
            logger.info(f"Found matching subscription {sub_id} (externalId.vendor: {external_vendor_id})")
            
            lines = sub.get('lines', [])
            if isinstance(lines, dict):
                lines = [lines]
            
            active_lines = [line for line in lines if line.get('status', '').lower() == 'active']
            if not active_lines:
                logger.warning(f"Subscription {sub_id} has no active lines, skipping")
                continue
            
            excel_items_map = {}
            for excel_item in excel_sub_data['items']:
                item_id = excel_item.get('Item ID')
                if item_id:
                    excel_items_map[item_id] = excel_item
            
            lines_to_update = []
            for line in active_lines:
                line_item_id = line.get('item', {}).get('id')
                if not line_item_id:
                    continue
                
                excel_item = excel_items_map.get(line_item_id)
                if not excel_item:
                    logger.debug(f"Subscription {sub_id} line with item {line_item_id} not found in Excel, skipping")
                    continue
                
                excel_markup = excel_item.get('Markup')
                if excel_markup is None:
                    logger.warning(f"Subscription {sub_id} line with item {line_item_id} has no markup in Excel, skipping")
                    continue
                
                line_id = line.get('id')
                if not line_id:
                    logger.error(f"Subscription {sub_id} line with item {line_item_id} has no line ID, skipping")
                    continue
                
                excel_unit_pp = excel_item.get('Unit PP')
                api_price = line.get('price', {})
                api_unit_pp = api_price.get('unitPP') if api_price else None
                
                final_unit_pp = None
                if args.keep_purchase_price:
                    if excel_unit_pp is not None and excel_unit_pp != 0.0:
                        final_unit_pp = excel_unit_pp
                    elif api_unit_pp is not None and api_unit_pp != 0.0:
                        final_unit_pp = api_unit_pp
                    
                    if final_unit_pp is None or final_unit_pp == 0.0:
                        logger.warning(
                            f"Subscription {sub_id} line with item {line_item_id} has no valid unitPP "
                            f"(Excel: {excel_unit_pp}, API: {api_unit_pp}). "
                            f"Will use markup instead for this line."
                        )
                        final_unit_pp = None
                else:
                    final_unit_pp = excel_unit_pp
                
                line_data = {
                    'line_id': line_id,
                    'markup': excel_markup,
                    'item_id': line_item_id,
                    'terms': line.get('terms', {}),
                    'quantity': line.get('quantity', 1),
                    'quantity_not_applicable': line.get('quantityNotApplicable', False),
                    'unit_pp': final_unit_pp,
                    'price': api_price
                }
                
                lines_to_update.append(line_data)
            
            if not lines_to_update:
                logger.warning(f"Subscription {sub_id} has no matching lines to update, skipping")
                continue
            
            logger.info(f"Updating {len(lines_to_update)} line(s) for subscription {sub_id}")
            
            success = update_subscription_markup(
                config['API_URL'],
                config['OPS_TOKEN'],
                sub_id,
                lines_to_update,
                logger,
                dry_run=dry_run,
                keep_purchase_price=args.keep_purchase_price
            )
            
            if success:
                updated_count += 1
                total_lines_updated += len(lines_to_update)
        
        logger.info("=" * 80)
        logger.info("SUMMARY")
        logger.info("=" * 80)
        logger.info(f"Total subscriptions in agreement: {len(agreement_subscriptions)}")
        logger.info(f"Total unique subscriptions in Excel: {len(excel_subscriptions)}")
        logger.info(f"Subscriptions found and matched: {found_count}")
        logger.info(f"Subscriptions updated: {updated_count}")
        logger.info(f"Total lines updated: {total_lines_updated}")
        logger.info(f"Subscriptions not found in Excel: {not_found_count}")
        if dry_run:
            logger.info("(Dry-run mode - no actual changes were made)")
        logger.info("=" * 80)
        
    except Exception as e:
        logger.error(f"Error in main execution: {str(e)}")
        raise


if __name__ == "__main__":
    main()

