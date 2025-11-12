import logging
import os
import sys
from pathlib import Path
from dotenv import load_dotenv
from datetime import datetime, timezone
from rich.logging import RichHandler
from urllib.parse import urljoin
from openpyxl import Workbook
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
from typing import Dict, List, Any, Optional
import argparse
import json

# Add the script's directory to the path to allow importing utils
script_dir = Path(__file__).parent
if str(script_dir) not in sys.path:
    sys.path.insert(0, str(script_dir))

from utils import (
    validate_agreement_id,
    validate_listing_id,
    validate_licensee_id,
    setup_logging,
    has_more_pages,
    validate_agreement_and_tokens,
    create_http_client,
    make_request_with_retry
)

# Load environment variables from ~/.mpt-clone-agreement
ENV_PATH = Path.home() / '.mpt-clone-agreement'
load_dotenv(ENV_PATH)


class ConfigurationManager:
    REQUIRED_VARS = {
        'OPS_TOKEN': 'API Operations Token',
        'VENDOR_TOKEN': 'Vendor API Token',
        'API_URL': 'API Base URL',
    }

    @classmethod
    def load_config(cls) -> Dict[str, str]:
        """Load and validate environment variables."""
        config = {key: os.getenv(key) for key in cls.REQUIRED_VARS}
        missing_vars = [f"{key} ({desc})" for key, desc in cls.REQUIRED_VARS.items() 
                       if not config.get(key)]
        
        if missing_vars:
            raise EnvironmentError(
                f"Missing required environment variables:\n"
                f"{chr(10).join(missing_vars)}\n"
                f"Please ensure these are set in {ENV_PATH}"
            )
        
        return config

def parse_arguments() -> argparse.Namespace:
    """
    Parse command line arguments.
    
    Returns:
        argparse.Namespace: Parsed command line arguments
    """
    parser = argparse.ArgumentParser(description='Dump Agreement Subscriptions Report')
    parser.add_argument(
        '--agreement-id',
        type=str,
        required=True,
        dest='agreement_id',
        help='Agreement ID to filter subscriptions (e.g., AGR-1234-5678-9012)'
    )
    listing_group = parser.add_mutually_exclusive_group(required=True)
    listing_group.add_argument(
        '--listing-id',
        type=str,
        dest='listing_id',
        help='Listing ID to use for cloning (e.g., LST-9279-6638)'
    )
    listing_group.add_argument(
        '--licensee-id',
        type=str,
        dest='licensee_id',
        help='Licensee ID to use for cloning (e.g., LIC-1234-5678-9012)'
    )
    parser.add_argument(
        '--debug',
        action='store_true',
        help='Enable debug level logging (more verbose output)'
    )
    args = parser.parse_args()
    
    # Validate argument formats
    validate_agreement_id(args.agreement_id)
    if args.listing_id:
        validate_listing_id(args.listing_id)
    elif args.licensee_id:
        validate_licensee_id(args.licensee_id)
    
    return args

class SubscriptionReport:
    def __init__(
        self,
        config: Dict[str, str],
        logger: logging.Logger,
        agreement_id: str
    ):
        self.config = config
        self.logger = logger
        self.agreement_id = agreement_id
        self.workbook: Optional[Workbook] = None
        self.worksheet = None
        self.problematic_subscriptions = []

    def create_workbook(self) -> None:
        """Initialize workbook and worksheet with headers."""
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        headers = [
            'ID', 'Vendor Sub ID', 'Client Sub ID', 'Name', 'Status', 'Agreement ID',
            'Agreement CCO', 'Agreement Client ID', 'Agreement Name', 'Agreement Vendor ID',
            'Agreement Authorization ID', 'Buyer ID', 'Buyer SCU', 'Buyer Name', 'Seller ID',
            'Seller Nav', 'Seller Name', 'Item Name', 'Item ID', 'Item MS ID', 'billing period',
            'Commitment period', 'Markup', 'Margin', 'Currency', 'Unit SP', 'Unit PP', 'Quantity',
            'AutoRenew', 'Start date', 'Commitment date', 'Original domain', 'From Migrated Data', 'Tier 2/Resell', 'MPN'
        ]
        for col, header in enumerate(headers, start=1):
            self.worksheet.cell(row=1, column=col, value=header)

    def safe_get(self, obj: Dict, *keys: str, default: Any = '') -> Any:
        """Safely get nested dictionary values."""
        current = obj
        try:
            for key in keys:
                if isinstance(current, list) and current:
                    current = current[0]
                current = current.get(key, default)
            return current if current is not None else default
        except (AttributeError, IndexError):
            return default

    def get_parameter_value(self, parameters: List[Dict], name: str) -> str:
        """Get parameter value by name from a list of parameters."""
        for param in parameters:
            if param.get('externalId').lower() == name.lower():
                return param.get('displayValue', '')
        return ''

    def get_mpn(self, certificates: List[Dict]) -> str:
        """Get MPN from certificates."""
        for cert in certificates:
            if cert.get('program').get('id','') == 'PRG-0742-8320':
                return cert.get('externalIds', {}).get('vendor', '-')
        return '-'

    def populate_worksheet(self, subscriptions: List[Dict[str, Any]]) -> None:
        """Populate worksheet with subscription data."""
        if not subscriptions:
            self.logger.error("No subscriptions data provided to populate_worksheet")
            return

        self.logger.info(f"Processing {len(subscriptions)} subscriptions")
        processed_count = 0
        skipped_count = 0
        row = 2  # Start from row 2 (row 1 is headers)

        for sub in subscriptions:
            try:
                self.logger.debug(f"Processing subscription {sub.get('id')}")
                if not sub.get('lines'):
                    self.problematic_subscriptions.append(sub)
                    self.logger.critical(f"Subscription {sub.get('id')} has no lines.")
                    skipped_count += 1
                    continue

                lines = sub.get('lines', [])
                if isinstance(lines, dict):
                    lines = [lines]
                    self.logger.debug(f"Converting single line to list for subscription {sub.get('id')}")

                active_lines = [line for line in lines if line.get('status', '').lower() == 'active']

                if not active_lines:
                    self.logger.warning(f"No active lines found for subscription {sub.get('id')}, skipping")
                    skipped_count += 1
                    continue

                if not sub.get('id') or not sub.get('name'):
                    self.logger.error(f"Missing basic subscription data for subscription {sub.get('id')}")
                    skipped_count += 1
                    continue

                for active_line in active_lines:
                    ms_sub = self.safe_get(sub, 'externalIds', 'vendor')
                    
                    self.logger.debug(f"Setting data for subscription {sub['id']} - Line item: {self.safe_get(active_line, 'item', 'id')}")

                    self.worksheet.cell(row=row, column=1, value=sub['id'])
                    self.worksheet.cell(row=row, column=2, value=ms_sub)
                    self.worksheet.cell(row=row, column=3, value=self.safe_get(sub, 'externalIds', 'client'))
                    self.worksheet.cell(row=row, column=4, value=sub['name'])
                    self.worksheet.cell(row=row, column=5, value=sub['status'])
                    self.worksheet.cell(row=row, column=6, value=self.safe_get(sub, 'agreement', 'id'))
                    self.worksheet.cell(row=row, column=7, value=self.safe_get(sub, 'agreement', 'externalIds', 'operations'))
                    self.worksheet.cell(row=row, column=8, value=self.safe_get(sub, 'agreement', 'externalIds', 'client'))
                    self.worksheet.cell(row=row, column=9, value=self.safe_get(sub, 'agreement', 'name'))
                    self.worksheet.cell(row=row, column=10, value=self.safe_get(sub, 'agreement', 'externalIds', 'vendor'))
                    self.worksheet.cell(row=row, column=11, value=self.safe_get(sub, 'agreement', 'authorization', 'externalIds', 'operations'))
                    self.worksheet.cell(row=row, column=12, value=self.safe_get(sub, 'buyer', 'externalIds', 'erpCustomer'))
                    self.worksheet.cell(row=row, column=13, value=self.safe_get(sub, 'buyer', 'id'))
                    self.worksheet.cell(row=row, column=14, value=self.safe_get(sub, 'buyer', 'name'))
                    self.worksheet.cell(row=row, column=15, value=self.safe_get(sub, 'seller', 'id'))
                    self.worksheet.cell(row=row, column=16, value=self.safe_get(sub, 'seller', 'externalId'))
                    self.worksheet.cell(row=row, column=17, value=self.safe_get(sub, 'seller', 'name'))
                    self.worksheet.cell(row=row, column=18, value=self.safe_get(active_line, 'item', 'name'))
                    self.worksheet.cell(row=row, column=19, value=self.safe_get(active_line, 'item', 'id'))
                    self.worksheet.cell(row=row, column=20, value=self.safe_get(active_line, 'item', 'externalIds', 'vendor'))
                    self.worksheet.cell(row=row, column=21, value=self.safe_get(sub, 'terms', 'period'))
                    self.worksheet.cell(row=row, column=22, value=self.safe_get(sub, 'terms', 'commitment'))

                    markup = round(self.safe_get(active_line, 'price', 'markup', default=0.0), 2)
                    if markup == 0.0:
                        markup = round(self.safe_get(sub, 'price', 'defaultMarkup', default=0.0), 2)
                    self.worksheet.cell(row=row, column=23, value=markup)

                    margin = round(self.safe_get(active_line, 'price', 'margin', default=0.0), 2)
                    if margin == 0.0:
                        default_markup = self.safe_get(sub, 'price', 'defaultMarkup', default=0.0)
                        margin = round((default_markup/100)/(1+(default_markup/100))*100, 2)
                    self.worksheet.cell(row=row, column=24, value=margin)

                    self.worksheet.cell(row=row, column=25, value=self.safe_get(active_line, 'price', 'currency'))
                    self.worksheet.cell(row=row, column=26, value=self.safe_get(active_line, 'price', 'unitSP'))
                    self.worksheet.cell(row=row, column=27, value=self.safe_get(active_line, 'price', 'unitPP'))
                    self.worksheet.cell(row=row, column=28, value=self.safe_get(active_line, 'quantity', default='1'))
                    self.worksheet.cell(row=row, column=29, value='Enabled' if sub.get('autoRenew') else 'Disabled')
                    self.worksheet.cell(row=row, column=30, value=sub.get('startDate', '').replace('T', ' ').replace('Z', ''))
                    self.worksheet.cell(row=row, column=31, value=sub.get('commitmentDate', '').replace('T', ' ').replace('Z', ''))
                    self.worksheet.cell(row=row, column=32, value=self.get_parameter_value(self.safe_get(sub, 'agreement', 'parameters', 'ordering', default=[]), "ExistingDomainName"))
                    self.worksheet.cell(row=row, column=33, value='No')
                    is_reseller = self.safe_get(sub, 'licensee', 'eligibility', 'partner', default=False)
                    mpn = '-'
                    if is_reseller:
                        mpn = self.get_mpn(sub.get('agreement', {}).get('certificates', []))
                    self.worksheet.cell(row=row, column=34, value=is_reseller)
                    self.worksheet.cell(row=row, column=35, value=mpn)

                    processed_count += 1
                    row += 1

            except Exception as e:
                self.logger.error(f"Error processing subscription {sub.get('id')}: {str(e)}")
                skipped_count += 1

        self.logger.info(f"Worksheet population complete:")
        self.logger.info(f"- Total subscriptions processed: {processed_count}")
        self.logger.info(f"- Total subscriptions skipped: {skipped_count}")
        self.logger.info(f"- Total rows in worksheet: {self.worksheet.max_row}")

    def set_column_widths(self) -> None:
        """Set appropriate column widths based on content."""
        if not self.worksheet:
            raise ValueError("Worksheet not initialized")

        dim_holder = DimensionHolder(worksheet=self.worksheet)
        
        for col in range(self.worksheet.min_column, self.worksheet.max_column + 1):
            width = 0
            for cell in self.worksheet[get_column_letter(col)]:
                if cell.value:
                    cell_length = len(str(cell.value))
                    width = max(width, cell_length)
            
            dim_holder[get_column_letter(col)] = ColumnDimension(
                self.worksheet, 
                min=col, 
                max=col, 
                width=min(width + 2, 50)  # Add padding, max width of 50
            )

        self.worksheet.column_dimensions = dim_holder

    def save_workbook(self, filename: str) -> None:
        """
        Save workbook locally.
        
        Args:
            filename: Path where to save the workbook
        """
        if not self.workbook:
            raise ValueError("Workbook not initialized")

        self.workbook.save(filename)
        self.logger.info(f"Saved report to {filename}")


    def get_subscriptions(self) -> List[Dict[str, Any]]:
        """Fetch subscriptions from API filtered by agreement ID."""
        items = []
        page = None
        limit = 1000
        offset = 0

        try:
            current_date = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%S.000Z")
        except Exception:
            current_date = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%S.000Z")
        
        url_base = urljoin(
            self.config['API_URL'],
            f'/public/v1/commerce/subscriptions?and(lt(audit.created.at,{current_date}),eq(agreement.id,{self.agreement_id}),eq(status,active))'
            '&select=agreement,lines,agreement.authorization.externalIds,agreement.listing.priceList,agreement.parameters, agreement.certificates,licensee,buyer,seller,audit,-agreement.subscriptions,-agreement.lines'
            f'&order=-audit.created.at'
        )
        
        with create_http_client(self.config['OPS_TOKEN'], 'Agreements Clone') as client:
            while has_more_pages(page):
                url = f"{url_base}&offset={offset}&limit={limit}"
                page = make_request_with_retry(
                    client=client,
                    method='GET',
                    url=url,
                    logger=self.logger,
                    parse_json=True
                )
                if not page:
                    self.logger.error(f"Failed to fetch subscriptions page at offset {offset}")
                    break
                
                page_content = page.get('data', [])
                items.extend(page_content)
                offset += limit
                self.logger.info(page.get('$meta', {}))

        return items

    def get_agreement(self) -> Optional[Dict[str, Any]]:
        """Fetch agreement details from API."""
        url = urljoin(
            self.config['API_URL'],
            f'/public/v1/commerce/agreements/{self.agreement_id}'
        )
        
        with create_http_client(self.config['OPS_TOKEN'], 'Agreements Clone') as client:
            agreement = make_request_with_retry(
                client=client,
                method='GET',
                url=url,
                logger=self.logger,
                parse_json=True
            )
        
        if agreement:
            self.logger.info(f"Successfully fetched agreement {self.agreement_id}")
        else:
            self.logger.error(f"Failed to fetch agreement {self.agreement_id}")
        
        return agreement
    
    def get_listing(self, listing_id: str) -> Optional[Dict[str, Any]]:
        """Fetch listing details from API."""
        url = urljoin(
            self.config['API_URL'],
            f'/public/v1/catalog/listings/{listing_id}?select=authorization,authorization.externalIds'
        )
        
        with create_http_client(self.config['OPS_TOKEN'], 'Agreements Clone') as client:
            listing = make_request_with_retry(
                client=client,
                method='GET',
                url=url,
                logger=self.logger,
                parse_json=True
            )
        
        if listing:
            self.logger.info(f"Successfully fetched listing {listing_id}")
        else:
            self.logger.error(f"Failed to fetch listing {listing_id}")
        
        return listing
    
    def get_licensee(
        self,
        licensee_id: str,
        seller_id: str,
        client_id: str
    ) -> Optional[Dict[str, Any]]:
        """
        Fetch licensee details from API.
        
        Args:
            licensee_id: Licensee ID to fetch
            seller_id: Seller ID for filtering
            client_id: Client ID for filtering
            
        Returns:
            Licensee object if found, None otherwise
        """
        url = urljoin(
            self.config['API_URL'],
            f'/public/v1/accounts/licensees?and(eq(id,{licensee_id}),eq(seller.id,{seller_id}),eq(account.id,{client_id}))'
        )
        
        with create_http_client(self.config['OPS_TOKEN'], 'Agreements Clone') as client:
            result = make_request_with_retry(
                client=client,
                method='GET',
                url=url,
                logger=self.logger,
                parse_json=True
            )
        
        if not result:
            self.logger.error(f"Failed to fetch licensee {licensee_id}")
            return None
        
        data = result.get('data', [])
        if len(data) == 0:
            self.logger.error(f"No licensee found with ID {licensee_id} for seller {seller_id} and client {client_id}")
            return None
        elif len(data) > 1:
            self.logger.error(f"Multiple licensees found ({len(data)}) with ID {licensee_id} for seller {seller_id} and client {client_id}")
            return None
        
        licensee = data[0]
        self.logger.info(f"Successfully fetched licensee {licensee_id}")
        return licensee
    
    def get_subscription_details(
        self,
        subscription_id: str
    ) -> Optional[Dict[str, Any]]:
        """Fetch subscription details from API using VENDOR_TOKEN."""
        url = urljoin(
            self.config['API_URL'],
            f'/public/v1/commerce/subscriptions/{subscription_id}'
        )
        
        with create_http_client(self.config['VENDOR_TOKEN'], 'Agreements Clone') as client:
            subscription = make_request_with_retry(
                client=client,
                method='GET',
                url=url,
                logger=self.logger,
                parse_json=True
            )
        
        if subscription:
            self.logger.debug(f"Successfully fetched subscription {subscription_id}")
        else:
            self.logger.error(f"Failed to fetch subscription {subscription_id}")
        
        return subscription

def main():
    args = parse_arguments()
    script_name = Path(__file__).stem
    logger = setup_logging(script_name, args.debug, args.agreement_id)
    
    try:
        logger.debug("Starting report generation in debug mode")
        config = ConfigurationManager.load_config()
        
        # Validate agreement and tokens before proceeding
        validate_agreement_and_tokens(
            args.agreement_id,
            config['API_URL'],
            config['OPS_TOKEN'],
            config['VENDOR_TOKEN'],
            logger
        )
        
        report = SubscriptionReport(config, logger, args.agreement_id)
        
        output_dir = Path("output") / args.agreement_id
        output_dir.mkdir(parents=True, exist_ok=True)
        
        logger.debug("Fetching agreement from API")
        agreement = report.get_agreement()
        if not agreement:
            logger.error(f"Could not fetch agreement {args.agreement_id}, stopping execution")
            return
        agreement_json_path = output_dir / 'agreement_object.json'
        with open(agreement_json_path, 'w', encoding='utf-8') as f:
            json.dump(agreement, f, indent=2, ensure_ascii=False)
        logger.info(f"Agreement saved to {agreement_json_path}")
        
        if args.listing_id:
            logger.debug("Fetching listing from API")
            listing = report.get_listing(args.listing_id)
            if listing:
                authorization = listing.get('authorization')
                if authorization:
                    authorization_json_path = output_dir / 'authorization.json'
                    with open(authorization_json_path, 'w', encoding='utf-8') as f:
                        json.dump(authorization, f, indent=2, ensure_ascii=False)
                    logger.info(f"Authorization saved to {authorization_json_path}")
                else:
                    logger.warning(f"No authorization found in listing {args.listing_id}")
                
                listing_auth_id = None
                try:
                    listing_auth_id = listing.get('authorization', {}).get('id')
                    if not listing_auth_id and isinstance(listing.get('authorization'), dict):
                        listing_auth_id = listing['authorization'].get('id')
                except Exception:
                    listing_auth_id = None
                
                if not listing_auth_id:
                    logger.error(f"Could not determine authorization.id from listing {args.listing_id}")
                else:
                    new_agreement = json.loads(json.dumps(agreement))
                    if 'id' in new_agreement:
                        del new_agreement['id']
                    if 'listing' not in new_agreement or not isinstance(new_agreement.get('listing'), dict):
                        new_agreement['listing'] = {}
                    new_agreement['listing']['id'] = args.listing_id
                    new_agreement['authorization'] = {'id': listing_auth_id}
                    new_path = output_dir / 'new_agreement_object.json'
                    with open(new_path, 'w', encoding='utf-8') as f:
                        json.dump(new_agreement, f, indent=2, ensure_ascii=False)
                    logger.info(f"New agreement object saved to {new_path}")
        
        elif args.licensee_id:
            logger.debug("Fetching licensees from API")
            
            source_licensee_id = agreement.get('licensee', {}).get('id')
            seller_id = agreement.get('seller', {}).get('id')
            client_id = agreement.get('client', {}).get('id')
            original_listing_id = agreement.get('listing', {}).get('id')
            
            if not source_licensee_id:
                logger.error("Could not find licensee.id in agreement")
                return
            if not seller_id:
                logger.error("Could not find seller.id in agreement")
                return
            if not client_id:
                logger.error("Could not find client.id in agreement")
                return
            if not original_listing_id:
                logger.error("Could not find listing.id in agreement")
                return
            
            logger.info(f"Source licensee ID: {source_licensee_id}")
            logger.info(f"Destination licensee ID: {args.licensee_id}")
            logger.info(f"Original listing ID: {original_listing_id}")
            logger.info(f"Seller ID: {seller_id}, Client ID: {client_id}")
            
            logger.debug(f"Fetching source licensee {source_licensee_id}")
            source_licensee = report.get_licensee(source_licensee_id, seller_id, client_id)
            if not source_licensee:
                logger.error(f"Failed to fetch source licensee {source_licensee_id}, stopping execution")
                return
            
            logger.debug(f"Fetching destination licensee {args.licensee_id}")
            destination_licensee = report.get_licensee(args.licensee_id, seller_id, client_id)
            if not destination_licensee:
                logger.error(f"Failed to fetch destination licensee {args.licensee_id}, stopping execution")
                return
            
            logger.debug(f"Fetching original listing {original_listing_id} to get authorization")
            original_listing = report.get_listing(original_listing_id)
            if not original_listing:
                logger.error(f"Failed to fetch original listing {original_listing_id}, stopping execution")
                return
            
            authorization = original_listing.get('authorization')
            if not authorization:
                logger.error(f"No authorization found in original listing {original_listing_id}")
                return
            
            authorization_json_path = output_dir / 'authorization.json'
            with open(authorization_json_path, 'w', encoding='utf-8') as f:
                json.dump(authorization, f, indent=2, ensure_ascii=False)
            logger.info(f"Authorization saved to {authorization_json_path}")
            
            listing_auth_id = authorization.get('id')
            if not listing_auth_id:
                logger.error(f"Could not determine authorization.id from original listing {original_listing_id}")
                return
            
            destination_buyer_id = destination_licensee.get('buyer', {}).get('id')
            if not destination_buyer_id:
                logger.error(f"Could not find buyer.id in destination licensee {args.licensee_id}")
                return
            
            logger.info(f"Destination licensee buyer ID: {destination_buyer_id}")
            
            new_agreement = json.loads(json.dumps(agreement))
            if 'id' in new_agreement:
                del new_agreement['id']
            new_agreement['licensee'] = {
                'id': args.licensee_id
            }
            new_agreement['buyer'] = {
                'id': destination_buyer_id
            }
            new_agreement['authorization'] = {'id': listing_auth_id}
            new_path = output_dir / 'new_agreement_object.json'
            with open(new_path, 'w', encoding='utf-8') as f:
                json.dump(new_agreement, f, indent=2, ensure_ascii=False)
            logger.info(f"New agreement object saved to {new_path}")
        
        logger.debug("Fetching subscriptions from API")
        subscriptions = report.get_subscriptions()
        
        subscriptions_without_vendor_id = []
        for sub in subscriptions:
            vendor_id = sub.get('externalIds', {}).get('vendor')
            if not vendor_id or (isinstance(vendor_id, str) and vendor_id.strip() == ''):
                subscriptions_without_vendor_id.append(sub)
        
        if len(subscriptions_without_vendor_id) > 1:
            logger.error(f"Found {len(subscriptions_without_vendor_id)} subscriptions with empty externalIds.vendor:")
            for sub in subscriptions_without_vendor_id:
                logger.error(f"  - Subscription {sub.get('id')} has no externalIds.vendor")
            logger.error("This will cause ambiguity when matching subscriptions. Stopping execution.")
            return
        elif len(subscriptions_without_vendor_id) == 1:
            sub_id = subscriptions_without_vendor_id[0].get('id')
            logger.warning(f"Found 1 subscription with empty externalIds.vendor: {sub_id}")
            logger.warning("This subscription may not be matchable in update_subscription_markups.py")

        logger.debug("Creating workbook and populating data")
        report.create_workbook()
        report.populate_worksheet(subscriptions)
        report.set_column_widths()

        # Save report to output folder
        filename = output_dir / 'subscriptions.xlsx'
        
        logger.debug(f"Saving workbook: {filename}")
        report.save_workbook(str(filename))
        
        logger.info(f"Report file saved at: {filename}")
        
        logger.debug("Dumping subscriptions as JSON files")
        dumped_count = 0
        for sub in subscriptions:
            sub_id = sub.get('id')
            if not sub_id:
                logger.warning(f"Skipping subscription with no ID: {sub}")
                continue
            
            logger.debug(f"Fetching full details for subscription {sub_id}")
            full_subscription = report.get_subscription_details(sub_id)
            if not full_subscription:
                logger.warning(f"Could not fetch subscription {sub_id}, skipping JSON dump")
                continue
            
            subscription_data = json.loads(json.dumps(full_subscription))
            if 'id' in subscription_data:
                del subscription_data['id']
            
            subscription_json_path = output_dir / f'{sub_id}.json'
            with open(subscription_json_path, 'w', encoding='utf-8') as f:
                json.dump(subscription_data, f, indent=2, ensure_ascii=False)
            
            dumped_count += 1
            logger.debug(f"Subscription {sub_id} saved to {subscription_json_path}")
        
        logger.info(f"Dumped {dumped_count} subscription(s) as JSON files")
        
        if report.problematic_subscriptions:
            logger.critical("=== Summary of All Problematic Subscriptions ===")
            for sub in report.problematic_subscriptions:
                logger.critical(f"Subscription ID: {sub.get('id', 'Unknown ID')}")
            logger.critical(f"Total problematic subscriptions: {len(report.problematic_subscriptions)}")
        
        logger.info("Report generation completed successfully")

    except Exception as e:
        logger.error(f"Error in main execution: {str(e)}")
        raise

if __name__ == "__main__":
    main()

